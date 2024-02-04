from openpyxl import Workbook
from docx import Document
from openpyxl import load_workbook

def create_empty_excel_file(filename):
    """创建一个空的Excel文件"""
    workbook = Workbook()
    workbook.save(filename)

def create_empty_word_file(filename):
    """创建一个空的Word文件"""
    document = Document()
    document.save(filename)



def update_excel_cell(filename, cell, value):
    """
    修改指定Excel文件中指定单元格的内容。

    :param filename: Excel文件的路径。
    :param cell: 要修改的单元格，例如 'A1'。
    :param value: 新值。
    """
    # 加载Excel文件
    workbook = load_workbook(filename)
    # 选择活动工作表
    sheet = workbook.active
    # 修改指定单元格的值
    sheet[cell] = value
    # 保存文件
    workbook.save(filename)

def merge_cells_in_excel(filename, cell_range):
    """
    合并指定Excel文件中的指定单元格范围。

    :param filename: Excel文件的路径。
    :param cell_range: 要合并的单元格范围，例如 'A1:D1'。
    """
    # 加载Excel文件
    workbook = load_workbook(filename)
    # 选择活动工作表
    sheet = workbook.active
    # 合并单元格
    sheet.merge_cells(cell_range)
    # 保存文件
    workbook.save(filename)

def get_cell_value(filename, cell):
    """
    获取指定Excel文件中指定单元格的内容。

    :param filename: Excel文件的路径。
    :param cell: 单元格位置，例如 'A1'。
    :return: 指定单元格的内容。
    """
    # 加载Excel文件
    workbook = load_workbook(filename)
    # 选择活动工作表
    sheet = workbook.active
    # 获取并返回指定单元格的值
    value = sheet[cell].value
    return value


